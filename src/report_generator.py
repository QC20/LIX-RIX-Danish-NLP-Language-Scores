"""
report_generator.py
--------------------
Generates two output formats from a list of DocumentResult objects:

1. A multi-sheet Excel workbook with colour coding, auto-widths, a chart
   of LIX scores, and a sheet of hardest sentences.

2. A self-contained HTML report that can be opened in any browser.

Usage (from main.py)
~~~~~~~~~~~~~~~~~~~~~
    from report_generator import generate_excel, generate_html

    generate_excel(results, output_path="report.xlsx")
    generate_html(results,  output_path="report.html")
"""

from __future__ import annotations

import html
import json
from datetime import datetime
from pathlib import Path

from text_analyzer import DocumentResult

# ---------------------------------------------------------------------------
# Colour palette (aligned with LIX bands)
# ---------------------------------------------------------------------------

_HEADER_FILL = "#2C3E50"
_HEADER_FONT = "#FFFFFF"
_ALT_ROW     = "#F2F3F4"

_SCORE_COLOURS = {
    "Very easy":           "#D5F5E3",
    "Easy":                "#A9DFBF",
    "Moderate":            "#FCF3CF",
    "Difficult":           "#FAD7A0",
    "Very difficult":      "#F5CBA7",
    "Extremely difficult": "#FADBD8",
}


# ===========================================================================
# Excel
# ===========================================================================

def generate_excel(results: list[DocumentResult], output_path: str | Path) -> Path:
    """
    Write a formatted Excel workbook to *output_path*.

    Sheets
    ------
    Summary         One row per document with all metrics, colour-coded by LIX.
    LIX Chart       Bar chart of LIX scores.
    Hard Sentences  Top difficult sentences across all documents.
    Top Long Words  Most frequent long words per document.
    Legend          Explanation of all columns and score bands.
    """
    try:
        import openpyxl
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel output. "
            "Install it with: pip install openpyxl"
        ) from exc

    output_path = Path(output_path)
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _build_summary_sheet(wb, results)
    _build_chart_sheet(wb, results)
    _build_sentences_sheet(wb, results)
    _build_words_sheet(wb, results)
    _build_legend_sheet(wb)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(wb, results: list[DocumentResult]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Summary")

    columns = [
        ("File",                      "file_name"),
        ("Language",                  "language"),
        ("LIX Score",                 "lix"),
        ("LIX Rating",                "lix_label"),
        ("Typical Text Type",         "lix_description"),
        ("RIX Score",                 "rix"),
        ("Grade Level (RIX)",         "rix_grade"),
        ("Word Count",                "word_count"),
        ("Sentence Count",            "sentence_count"),
        ("Avg Words / Sentence",      "avg_words_per_sentence"),
        ("Avg Word Length (chars)",   "avg_word_length"),
        ("Long Words (LIX, count)",   "long_word_count"),
        ("Long Words (%)",            "long_word_ratio"),
        ("Unique Content Words",      "unique_word_count"),
        ("Type-Token Ratio",          "type_token_ratio"),
        ("Nouns",                     "noun_count"),
        ("Verbs",                     "verb_count"),
        ("Adjectives",                "adjective_count"),
    ]

    # Header row
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, result in enumerate(results, start=2):
        fill_colour = _SCORE_COLOURS.get(result.lix_label, "#FFFFFF")
        alt_fill = PatternFill("solid", fgColor=fill_colour.lstrip("#"))
        plain_fill = PatternFill("solid", fgColor="F2F3F4")

        for col_idx, (_, attr) in enumerate(columns, start=1):
            value = getattr(result, attr)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Colour the LIX score and label columns; alternate grey for rest
            if attr in ("lix", "lix_label", "lix_description"):
                cell.fill = alt_fill
            elif row_idx % 2 == 0:
                cell.fill = plain_fill

    # Auto-width (capped at 40)
    for col_idx in range(1, len(columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(results) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "C2"


def _build_chart_sheet(wb, results: list[DocumentResult]) -> None:
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("LIX Chart")

    # Write data for the chart
    ws.cell(row=1, column=1, value="File").font = Font(bold=True)
    ws.cell(row=1, column=2, value="LIX Score").font = Font(bold=True)

    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r.file_name)
        ws.cell(row=i, column=2, value=r.lix)

    # Reference lines (written as separate series so legend labels work)
    band_row = len(results) + 4
    ws.cell(row=band_row, column=1, value="Reference thresholds").font = Font(bold=True)
    for offset, (threshold, label) in enumerate(
        [(20, "Very easy <20"), (30, "Easy <30"), (40, "Moderate <40"),
         (50, "Difficult <50"), (60, "Very difficult <60")],
        start=1
    ):
        ws.cell(row=band_row + offset, column=1, value=label)
        ws.cell(row=band_row + offset, column=2, value=threshold)

    # Build chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "LIX Readability Scores"
    chart.y_axis.title = "LIX Score"
    chart.x_axis.title = "Document"
    chart.style = 10
    chart.width = 24
    chart.height = 14

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(results) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(results) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "D2")


def _build_sentences_sheet(wb, results: list[DocumentResult]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Hard Sentences")

    headers = ["File", "Sentence", "Words", "Long Words", "Long Word %"]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    row = 2
    for result in results:
        for sent in result.hardest_sentences:
            ws.cell(row=row, column=1, value=result.file_name)
            cell = ws.cell(row=row, column=2, value=sent.text)
            cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=3, value=sent.word_count)
            ws.cell(row=row, column=4, value=sent.long_word_count)
            ws.cell(row=row, column=5, value=f"{sent.lix_contribution:.1f}%")
            row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    for col in ["C", "D", "E"]:
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = "C2"


def _build_words_sheet(wb, results: list[DocumentResult]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Top Long Words")

    headers = ["File", "Word", "Occurrences"]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    row = 2
    for result in results:
        for word, count in result.top_long_words:
            ws.cell(row=row, column=1, value=result.file_name)
            ws.cell(row=row, column=2, value=word)
            ws.cell(row=row, column=3, value=count)
            row += 1

    for col, width in [("A", 22), ("B", 28), ("C", 14)]:
        ws.column_dimensions[col].width = width


def _build_legend_sheet(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Legend")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    header_fill = PatternFill("solid", fgColor="2C3E50")

    def heading(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def row_pair(row, label, value):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True)

    r = 1
    heading(r, "LIX Score Bands")
    r += 1
    bands = [
        ("<20",  "Very easy",           "Children's books, simple instructions"),
        ("20-30","Easy",                "Fiction, popular non-fiction"),
        ("30-40","Moderate",            "Newspapers, general magazines"),
        ("40-50","Difficult",           "Academic papers, technical reports"),
        ("50-60","Very difficult",       "Scientific literature, legal texts"),
        (">60",  "Extremely difficult", "Specialist research, regulatory documents"),
    ]
    for score, label, desc in bands:
        row_pair(r, f"LIX {score}  ({label})", desc)
        r += 1

    r += 1
    heading(r, "Column Descriptions")
    r += 1
    col_descs = [
        ("LIX Score",              "Björnsson (1968). LIX = avg_words_per_sentence + long_words%. Target ≤40 for public-facing text."),
        ("RIX Score",              "Anderson (1983). RIX = long_words / sentences. Maps directly to school grade levels."),
        ("Grade Level (RIX)",      "Estimated school grade required to read the text comfortably."),
        ("Avg Words / Sentence",   "Mean sentence length. Shorter sentences are generally easier to follow."),
        ("Avg Word Length (chars)","Mean number of characters per word. Higher values correlate with greater difficulty."),
        ("Long Words (LIX, count)","Words exceeding 6 characters (Scandinavian threshold for LIX)."),
        ("Long Words (%)",         "Percentage of all words that are 'long' by the LIX definition."),
        ("Unique Content Words",   "Distinct words after removing stop words."),
        ("Type-Token Ratio",       "Unique content words divided by total content words. Higher = richer vocabulary."),
        ("Nouns / Verbs / Adj.",   "Part-of-speech counts from the spaCy language model."),
    ]
    for label, desc in col_descs:
        row_pair(r, label, desc)
        r += 1

    r += 1
    heading(r, "Formula Reference")
    r += 1
    row_pair(r, "LIX", "(words / sentences) + (long_words_6+ × 100 / words)")
    r += 1
    row_pair(r, "RIX", "long_words_7+ / sentences")
    r += 1

    r += 1
    heading(r, "Recommendation for Public Writing")
    r += 1
    row_pair(r, "Target LIX", "≤40 for newspapers; ≤30 for general audiences")
    r += 1
    row_pair(r, "Target RIX grade", "≤8 for public-facing documents")


# ===========================================================================
# HTML
# ===========================================================================

def generate_html(results: list[DocumentResult], output_path: str | Path) -> Path:
    """Write a self-contained HTML report to *output_path*."""
    output_path = Path(output_path)
    content = _render_html(results)
    output_path.write_text(content, encoding="utf-8")
    return output_path


def _render_html(results: list[DocumentResult]) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    rows = ""
    for r in results:
        bg = r.lix_hex
        rows += f"""
        <tr>
          <td>{html.escape(r.file_name)}</td>
          <td>{html.escape(r.language)}</td>
          <td style="background:{bg};color:#fff;font-weight:bold">{r.lix:.2f}</td>
          <td style="background:{bg};color:#fff">{html.escape(r.lix_label)}</td>
          <td>{r.rix:.2f}</td>
          <td>{html.escape(str(r.rix_grade))}</td>
          <td>{r.word_count}</td>
          <td>{r.sentence_count}</td>
          <td>{r.avg_words_per_sentence}</td>
          <td>{r.avg_word_length}</td>
          <td>{r.long_word_ratio}%</td>
          <td>{r.type_token_ratio:.3f}</td>
        </tr>"""

    hard_rows = ""
    for r in results:
        for s in r.hardest_sentences:
            escaped = html.escape(s.text)
            hard_rows += f"""
            <tr>
              <td>{html.escape(r.file_name)}</td>
              <td class="sentence-cell">{escaped}</td>
              <td>{s.word_count}</td>
              <td>{s.lix_contribution:.1f}%</td>
            </tr>"""

    word_rows = ""
    for r in results:
        for word, count in r.top_long_words[:10]:
            word_rows += f"""
            <tr>
              <td>{html.escape(r.file_name)}</td>
              <td>{html.escape(word)}</td>
              <td>{count}</td>
            </tr>"""

    lix_json = json.dumps([
        {"file": r.file_name, "lix": r.lix, "color": r.lix_hex}
        for r in results
    ])

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Readability Report</title>
<style>
  :root {{
    --bg: #F8F9FA; --surface: #FFFFFF; --border: #DEE2E6;
    --header: #2C3E50; --header-fg: #FFFFFF; --accent: #2980B9;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: system-ui, sans-serif; background: var(--bg);
          color: #2C3E50; line-height: 1.6; padding: 2rem; }}
  h1 {{ font-size: 1.8rem; margin-bottom: 0.25rem; color: var(--header); }}
  .meta {{ color: #7F8C8D; font-size: 0.9rem; margin-bottom: 2rem; }}
  h2 {{ font-size: 1.2rem; margin: 2rem 0 0.75rem; color: var(--header);
        border-bottom: 2px solid var(--accent); padding-bottom: 0.25rem; }}
  .table-wrap {{ overflow-x: auto; border-radius: 6px;
                 box-shadow: 0 1px 4px rgba(0,0,0,.12); margin-bottom: 1.5rem; }}
  table {{ border-collapse: collapse; width: 100%; background: var(--surface); }}
  th {{ background: var(--header); color: var(--header-fg); padding: 10px 12px;
        text-align: left; white-space: nowrap; font-size: 0.85rem; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid var(--border);
        font-size: 0.85rem; vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:nth-child(even) td {{ background: #F2F3F4; }}
  .sentence-cell {{ max-width: 500px; }}
  .band-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
                gap: 1rem; margin-bottom: 1.5rem; }}
  .band-card {{ border-radius: 8px; padding: 1rem; color: #fff; }}
  .band-card strong {{ display: block; font-size: 1.1rem; }}
  .band-card span {{ font-size: 0.8rem; opacity: .9; }}
  canvas {{ max-width: 100%; }}
</style>
</head>
<body>
<h1>Readability Report</h1>
<p class="meta">Generated {now} &bull; {len(results)} document(s)</p>

<h2>Score Reference</h2>
<div class="band-grid">
  <div class="band-card" style="background:#27AE60"><strong>LIX &lt;20</strong><span>Very easy &mdash; children's books</span></div>
  <div class="band-card" style="background:#2ECC71"><strong>LIX 20-30</strong><span>Easy &mdash; fiction, popular non-fiction</span></div>
  <div class="band-card" style="background:#F39C12"><strong>LIX 30-40</strong><span>Moderate &mdash; newspapers</span></div>
  <div class="band-card" style="background:#E67E22"><strong>LIX 40-50</strong><span>Difficult &mdash; academic texts</span></div>
  <div class="band-card" style="background:#E74C3C"><strong>LIX 50-60</strong><span>Very difficult &mdash; scientific</span></div>
  <div class="band-card" style="background:#922B21"><strong>LIX &gt;60</strong><span>Extremely difficult &mdash; specialist</span></div>
</div>

<h2>Document Summary</h2>
<div class="table-wrap"><table>
  <thead><tr>
    <th>File</th><th>Lang</th><th>LIX</th><th>Rating</th>
    <th>RIX</th><th>Grade</th><th>Words</th><th>Sentences</th>
    <th>Avg Words/Sent</th><th>Avg Word Len</th><th>Long Words %</th><th>TTR</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table></div>

<h2>Hardest Sentences</h2>
<div class="table-wrap"><table>
  <thead><tr><th>File</th><th>Sentence</th><th>Words</th><th>Long Word %</th></tr></thead>
  <tbody>{hard_rows}</tbody>
</table></div>

<h2>Most Frequent Long Words</h2>
<div class="table-wrap"><table>
  <thead><tr><th>File</th><th>Word</th><th>Occurrences</th></tr></thead>
  <tbody>{word_rows}</tbody>
</table></div>

<h2>LIX Score Chart</h2>
<canvas id="chart" height="80"></canvas>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>
const data = {lix_json};
new Chart(document.getElementById("chart"), {{
  type: "bar",
  data: {{
    labels: data.map(d => d.file),
    datasets: [{{
      label: "LIX Score",
      data: data.map(d => d.lix),
      backgroundColor: data.map(d => d.color),
      borderRadius: 4,
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => "LIX: " + ctx.parsed.y.toFixed(2) }} }}
    }},
    scales: {{
      y: {{
        beginAtZero: true,
        title: {{ display: true, text: "LIX Score" }},
        grid: {{ color: "#DEE2E6" }}
      }}
    }}
  }}
}});
</script>
</body>
</html>
"""